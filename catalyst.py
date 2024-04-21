import math
import openpyxl as xl

def distance(s_lat, s_lon, d_lat, d_lon):
    earth_radius = 3959.0

    s_lat = s_lat*math.pi/180.0
    s_lon = math.radians(s_lon)
    d_lat = math.radians(d_lat)
    d_lon = math.radians(d_lon)

    dist = math.sin((d_lat - s_lat)/2)**2 + math.cos(s_lat)*math.cos(d_lat) * math.sin((d_lon - s_lon)/2)**2

    return 2 * earth_radius * math.asin(math.sqrt(dist))


zip_info = []
med_household_income = 84385
low_income_limit = 54300
income_database = xl.load_workbook('zip_code_income_edit.xlsx')
income_sheet = income_database['Sheet1']
'''with open('zip_code_income_edit.txt', 'r', encoding='utf-8') as file:
    income_database = file.readlines()
for zip in income_database:
    num, income, ignore, zipcode, ignore, lat, lon = income_database.strip().split('    ')
    zip_info.append(zipcode, income, lat, lon)'''
zipincome_col = 2
zipcode2_col = 5
row = 1
zip2lat_col = 6
zip2lon_col = 7

MA_zipcodes2 = []
for row in range(2, 523):
        MA_zipcode2 = ['0' + str(income_sheet.cell(row, zipcode2_col).value),
                      income_sheet.cell(row, zip2lat_col).value,
                      income_sheet.cell(row, zip2lon_col).value, income_sheet.cell(row, zipincome_col).value]
        MA_zipcodes2.append(MA_zipcode2)

print(str(len(zip_info)))

# get facilities with emission greater than 1,000
emission_data = xl.load_workbook('ghgp_data_2021.xlsx')
emission_sheet = emission_data['Direct Emitters']
emission_facility_id_col = 1
emission_facility_name_col = 3
emission_state_col = 5
emission_lat_col = 9
emission_lon_col = 10
emission_amount_col = 14
emission_limit = 40000

MA_emission_sources = []
for row in range(5, emission_sheet.max_row + 1):
    if emission_sheet.cell(row, emission_state_col).value == 'MA' and \
            emission_sheet.cell(row, emission_amount_col).value >= emission_limit:
        MA_emission_source = [emission_sheet.cell(row, emission_facility_id_col).value,
                              emission_sheet.cell(row, emission_facility_name_col).value,
                              emission_sheet.cell(row, emission_lat_col).value,
                              emission_sheet.cell(row, emission_lon_col).value,
                              emission_sheet.cell(row, emission_amount_col).value]#(facility ID, facility name, latitude, longitude, emission amount)
        MA_emission_sources.append(MA_emission_source)
        
print(len(MA_emission_sources))

zipcodes_low_income = []
for zip in zip_info:
    if zip[3] < low_income_limit:
        zipcodes_low_income.append([zip[0], zip[3]])

print(len(zipcodes_low_income))

radiusemissions = []
radius = 0
radius_step = 1
radius_last = 5
zip_high_emissions = []
for radius in range(radius, radius_last):
    radius = radius + radius_step 
    for item1 in zip_info:
        total_emission = 0.0
        for item2 in MA_emission_sources:
            d = distance(item1[1], item1[2], item2[2], item2[3])
            if d < radius:
                total_emission += item2[4]
                if total_emission > 0.0:
                    zip_high_emissions.append([radius, item1[0], total_emission, item1[3], item1[1], item1[2]])
        radiusemissions.append([radius, item1[0], total_emission, item1[3], item1[1], item1[2]])

radius_high_emissions_zip = []
zipcodes_high_emissions = []
for radius in range(radius, radius_last):
   radius = radius + radius_step
   for item1 in zip_info:
       total_emission = 0.0
       for item2 in MA_emission_sources:
           d = distance(item1[1], item1[2], item2[2], item2[3])
           if d < radius:
               total_emission += item2[4]
               print('Facility name: ' + item2[1] + '           Emission: ' + str(item2[4]))
       if total_emission > 0.0:
           zipcodes_high_emissions.append([radius, item1[0], total_emission, item1[3], item1[1], item1[2]])
           print('Zipcode: ' + item1[0] + '           Total emission: ' + str(total_emission) + '             Median Household Income: $' + str(item1[3]))
           print('***********************************************************')
   print('Total number of MA zipcodes within ' + str(radius) + ' miles of at least 1,000 tons of direct greenhouse gas emission: ' +
       str(len(zipcodes_high_emissions)))
   print('***********************************************************')
   print('***********************************************************')
   print('***********************************************************')
   print('***********************************************************')
   print('***********************************************************')
   total_med_income = 0
   for zip in zipcodes_high_emissions:
       total_med_income = zip[2]
   radius_high_emissions_zip.append([radius, len(zipcodes_high_emissions), total_med_income/len(zipcodes_high_emissions)])
print(radius_high_emissions_zip)